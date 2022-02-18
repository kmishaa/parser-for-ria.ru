import requests
from bs4 import BeautifulSoup
import openpyxl
import os

URL = 'https://rsport.ria.ru/services/search/getmore/?query=0&offset=0&list_sids[]=rsport&project[]=rsport&interval=period&date_from=2018-02-09&date_to=2018-02-25'

HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.85 YaBrowser/21.11.1.932 Yowser/2.5 Safari/537.36',
               'accept': '*/*'}

FILE = 'results_ria.xlsx'

def get_html(url):
    r = requests.get(url, headers=HEADERS) #получения html-кода с сайта
    return r

def get_all_articles(url): #получение всех статей с полностью прокрученной страницы
    articles = []
    counter = 0
    offset = url.split('&offset=')[1].split('&list_sids')[0]
    while True:
        print (f'Запрос {counter // 20 + 1}')
        link = url.replace(f'&offset={offset}&list_sids', f'&offset={counter}&list_sids') #составление очередной ссылки
        articles_20 = get_content((get_html(link)).text) #получение 20-ти новых статей
        if (len(articles_20) == 0): #проверка на конец страницы
            break
        counter += 20
        articles.extend(articles_20)

    return articles

def get_text(link): #получение текста статьи
    text = ''
    html = get_html(link)
    soup = BeautifulSoup(html.text, 'html.parser')
    all_texts = soup.find('div', class_='article__body js-mediator-article mia-analytics')
    if all_texts: #проверка на наличие текста статьи
        all_p = all_texts.find_all('p') #получение всех абзацев
        all_p.extend(all_texts.find_all('strong'))
        all_p.extend(all_texts.find_all('div', class_='article__text'))
        for p in all_p:
            text += p.get_text() + ' '
    else:
        text = 'NO TEXT'
    text = text.replace('\n', ' ') #удаление переносов строк
    
    return text

def get_content(html): #получение 20-ти статей с сайта
    articles = []
    soup = BeautifulSoup(html, 'html.parser')
    all_articles = soup.find_all('div', class_='list-item')
    counter = 1
    for article in all_articles:
        #print (f'Обработка статьи {counter} из {len(all_articles)}')
        data = article.find('a', class_='list-item__title color-font-hover-only')
        title = data.get_text() #получение заголовка
        link = data.get('href') #получение ссылки на текст статьи
        date = article.find('div', class_='list-item__date').get_text() #получение даты публикации статьи
        articles.append({
            'title': title.strip(),
            'date': date,
            'text': get_text(link)
        })
        counter += 1
        
    return articles

def save_file(items, path): #сохранение данных в файл
    if (os.path.exists(path)): #проверка на существование файла
        book = openpyxl.load_workbook(path)
        sheet = book.active
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'] = 'Название статьи'
        sheet['B1'] = 'Дата'
        sheet['C1'] = 'Текст статьи'

    row = sheet.max_row + 1 #дописываем в конец файла

    for item in items:
        title = item['title']
        date = item['date']
        text = item['text']
        sheet[row][0].value = title
        sheet[row][1].value = date
        sheet[row][2].value = text
        row += 1

    book.save(path)
    book.close()

def make_link(date_1, date_2): #составление ссылки по запросу
    link = URL
    dateFrom = link.split('&date_from=')[1].split('&date_to=')[0]
    dateTo = link.split('&date_to=')[1]
    link = link.replace(f'&date_from={dateFrom}&date_to=', f'&date_from=' + date_1 + '&date_to=')
    link = link.replace(f'&date_to={dateTo}', f'&date_to=' + date_2)
    return link

def parse():
    dates = ['2018-02-09', '2018-02-25', '2018-05-09', '2018-05-25', '2021-07-23', '2021-08-08', '2021-10-23', '2021-11-08'] #все исследуемые даты
    prints = { #выводы
        'Подсчет количества статей в период с 9 февраля по 25 февраля 2018 года..': make_link(dates[0], dates[1]),
        'Подсчет количества статей в период с 9 мая по 25 мая 2018 года..': make_link(dates[2], dates[3]),
        'Подсчет количества статей в период с 23 июля по 8 августа 2021 года..': make_link(dates[4], dates[5]),
        'Подсчет количества статей в период с 23 октября по 8 ноября 2021 года..': make_link(dates[6], dates[7]),
        'Подсчет количества статей 9 февраля 2018 года..': make_link(dates[0], dates[0]),
        'Подсчет количества статей 25 февраля 2018 года..': make_link(dates[1], dates[1]),
        'Подсчет количества статей 23 июля 2021 года..': make_link(dates[4], dates[4]),
        'Подсчет количества статей 8 августа 2021 года..': make_link(dates[5], dates[5])
    }
    
    html = get_html(URL)
    if (html.status_code == 200): #проверка доступа к сайту
        all_articles = []
        for printer, link in prints.items():
            print (printer)
            articles = get_all_articles(link) #получение списка статей по очередному запросу
            all_articles.extend(articles)
            save_file(articles, FILE) #сохранение
            print (f'Найдено {len(articles)} статей\n\n')

        print ('Подсчет количества статей со словосочетанием "церемония открытия" или "церемония закрытия"..')
        articles = all_articles
        ceremony = 0
        medvedeva = 0
        zagitova = 0
        averina = 0
        romashina = 0
        counter = 1
        for article in articles: #поиск по ключевым словам
            print (f'Исследуется статья номер {counter} из {len(articles)}')
            if ('еремони' in article['text'] and ('открыти' in article['text'] or 'закрыти' in article['text'])):
                find = [article]
                save_file(find, FILE)
                ceremony += 1
            if ('Евгени' in article['text'] and 'Медведев' in article['text']):
                find = [article]
                save_file(find, FILE)
                medvedeva += 1
            if ('Алин' in article['text'] and 'Загитов' in article['text']):
                find = [article]
                save_file(find, FILE)
                zagitova += 1
            if ('Дин' in article['text'] and 'Аверин' in article['text']):
                find = [article]
                save_file(find, FILE)
                averina += 1
            if ('Светлан' in article['text'] and 'Ромашин' in article['text']):
                find = [article]
                save_file(find, FILE)
                romashina += 1
            counter += 1
        print (f'Найдено {ceremony} статей\n\n')

        print ('Подсчет количества статей со словосочетанием "Евгения Медведева"..')
        print (f'Найдено {medvedeva} статей\n\n')

        print ('Подсчет количества статей со словосочетанием "Алина Загитова"..')
        print (f'Найдено {zagitova} статей\n\n')
        
        print ('Подсчет количества статей со словосочетанием "Дина Аверина"..')
        print (f'Найдено {averina} статей\n\n')

        print ('Подсчет количества статей со словосочетанием "Светлана Ромашина"..')
        print (f'Найдено {romashina} статей\n\n')
        
parse()
